#!/usr/bin/env python3
"""
CITS3200 Team Leader Spreadsheet Automation

Minimalistic overhaul:
- Structured logging with Rich
- Optional config via YAML and environment variables
- Exact, case-insensitive name matching
- Template validation up front
- Dry-run mode to skip writes/uploads
- Uses pathlib for paths
- Uses `rclone mkdir` for folder creation
"""

from __future__ import annotations

import sys
import subprocess
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook
import shutil
import zipfile
import logging
from rich.logging import RichHandler
import os
import yaml
from dotenv import load_dotenv


# ---------- Logging ----------
load_dotenv()
logging.basicConfig(
    level=logging.INFO,
    format="%(message)s",
    handlers=[RichHandler(rich_tracebacks=True, markup=True)],
)
logger = logging.getLogger("cits3200")


# ---------- Configuration ----------
@dataclass
class AppConfig:
    remote_name: str = "gdrive"
    remote_base_path: str = "CITS3200"
    master_file: str = "CITS3200_Group_31.xlsx"
    group_template: str = "templates/TimeSheet_GroupX_WkY.xlsx"
    individual_template: str = "templates/Booked_Hours_YourName.xlsx"
    output_dir: str = "output"
    team_members: List[str] = field(
        default_factory=lambda: [
            "Steve",
            "Ben",
            "Dongkai",
            "William",
            "Jeremy",
            "Noah",
        ]
    )
    group_number: int = 31
    # Optional secondary upload target (e.g., SharePoint/Teams)
    secondary_remote_name: Optional[str] = "teams"
    secondary_remote_base_path: Optional[str] = "General/Group_31"
    # Optional rclone flags for secondary remote copy operations (mitigate OneDrive hash/size mismatch)
    secondary_remote_copy_flags: List[str] = field(default_factory=lambda: [
        "--ignore-size",
        "--ignore-checksum",
        "--retries", "5",
        "--low-level-retries", "10",
        "--transfers", "1",
    ])


def load_config(config_path: Optional[Path]) -> AppConfig:
    cfg = AppConfig()

    # YAML config (optional)
    if config_path and config_path.exists():
        try:
            data = yaml.safe_load(config_path.read_text()) or {}
            cfg.remote_name = data.get("remote_name", cfg.remote_name)
            cfg.remote_base_path = data.get("remote_base_path", cfg.remote_base_path)
            cfg.master_file = data.get("master_file", cfg.master_file)
            cfg.group_template = data.get("group_template", cfg.group_template)
            cfg.individual_template = data.get("individual_template", cfg.individual_template)
            cfg.output_dir = data.get("output_dir", cfg.output_dir)
            cfg.team_members = data.get("team_members", cfg.team_members)
            cfg.group_number = int(data.get("group_number", cfg.group_number))
            cfg.secondary_remote_name = data.get("secondary_remote_name", cfg.secondary_remote_name)
            cfg.secondary_remote_base_path = data.get("secondary_remote_base_path", cfg.secondary_remote_base_path)
            cfg.secondary_remote_copy_flags = data.get("secondary_remote_copy_flags", cfg.secondary_remote_copy_flags)
        except Exception as e:
            logger.warning(f"[yellow]⚠️  Failed to read config file:[/yellow] {e}")

    # Environment overrides (optional)
    cfg.remote_name = os.getenv("REMOTE_NAME", cfg.remote_name)
    cfg.remote_base_path = os.getenv("REMOTE_BASE_PATH", cfg.remote_base_path)
    cfg.master_file = os.getenv("MASTER_FILE", cfg.master_file)
    cfg.group_template = os.getenv("GROUP_TEMPLATE", cfg.group_template)
    cfg.individual_template = os.getenv("INDIVIDUAL_TEMPLATE", cfg.individual_template)
    cfg.output_dir = os.getenv("OUTPUT_DIR", cfg.output_dir)
    cfg.group_number = int(os.getenv("GROUP_NUMBER", cfg.group_number))
    cfg.secondary_remote_name = os.getenv("SECONDARY_REMOTE_NAME", cfg.secondary_remote_name)
    cfg.secondary_remote_base_path = os.getenv("SECONDARY_REMOTE_BASE_PATH", cfg.secondary_remote_base_path)
    cfg.secondary_remote_copy_flags = os.getenv("SECONDARY_REMOTE_COPY_FLAGS", cfg.secondary_remote_copy_flags)

    team_members_env = os.getenv("TEAM_MEMBERS")
    if team_members_env:
        cfg.team_members = [m.strip() for m in team_members_env.split(",") if m.strip()]

    return cfg


class CITS3200Automation:
    def __init__(self, config: AppConfig, week_number: Optional[int] = None, dry_run: bool = False, yes_all: bool = False):
        self.config = config
        self.week_number: Optional[int] = week_number
        self.dry_run = dry_run
        self.yes_all = yes_all

        self.master_file_path = Path(self.config.master_file)
        self.group_template_path = Path(self.config.group_template)
        self.individual_template_path = Path(self.config.individual_template)
        self.output_dir = Path(self.config.output_dir)

        # Ensure output directory exists (skip in dry-run)
        if not self.dry_run:
            self.output_dir.mkdir(parents=True, exist_ok=True)

    def clear_output_dir(self) -> None:
        """Clear all contents from the output directory."""
        if not self.output_dir.exists():
            return
        if self.dry_run:
            logger.info(f"[cyan]DRY-RUN[/cyan] Would clear directory: {self.output_dir}")
            for path in sorted(self.output_dir.iterdir()):
                logger.info(f"[cyan]DRY-RUN[/cyan] Would remove: {path}")
            return
        try:
            for path in self.output_dir.iterdir():
                if path.is_file() or path.is_symlink():
                    path.unlink(missing_ok=True)
                elif path.is_dir():
                    shutil.rmtree(path, ignore_errors=True)
            logger.info(f"🧹 Cleared output directory: {self.output_dir}")
        except Exception as e:
            logger.error(f"❌ Failed to clear output directory '{self.output_dir}': {e}")
            raise

    def prompt_for_week(self) -> bool:
        """Prompt user for week number"""
        while True:
            try:
                week = int(input("Enter week number (2-11): "))
                if 2 <= week <= 11:
                    self.week_number = week
                    logger.info(f"✅ Week {week} selected")
                    return True
                else:
                    logger.error("❌ Week number must be between 2 and 11")
            except ValueError:
                logger.error("❌ Please enter a valid integer")

    def copy_master_from_gdrive(self) -> bool:
        """Copy the latest master file from Google Drive"""
        logger.info("📥 Copying master file from Google Drive...")
        if self.dry_run:
            logger.info("[cyan]DRY-RUN[/cyan] Would run: rclone copy "
                        f"{self.config.remote_name}:{self.config.remote_base_path}/{self.config.master_file} ./")
            return True
        try:
            result = subprocess.run([
                "rclone", "copy",
                f"{self.config.remote_name}:{self.config.remote_base_path}/{self.config.master_file}",
                str(Path("."))
            ], capture_output=True, text=True)

            if result.returncode == 0:
                logger.info("✅ Master file copied successfully")
            else:
                logger.error(f"❌ Error copying master file: {result.stderr}")
                return False
        except Exception as e:
            logger.error(f"❌ Error accessing Google Drive: {e}")
            return False
        return True

    def create_group_timesheet(self) -> Optional[Path]:
        """Create the group timesheet from template"""
        logger.info("📊 Creating group timesheet...")

        # Compute output path
        output_file = self.output_dir / f"TimeSheet_Group{self.config.group_number}_Wk{self.week_number}.xlsx"

        if self.dry_run:
            logger.info(f"[cyan]DRY-RUN[/cyan] Would copy group template → {output_file}")
            return output_file

        # Copy template with proper naming
        shutil.copy2(self.group_template_path, output_file)

        # Load both files
        master_wb = load_workbook(self.master_file_path, data_only=True)
        output_wb = load_workbook(output_file)

        # Transfer data from master to template
        self._transfer_group_data(master_wb, output_wb)

        # Save the output file
        output_wb.save(output_file)
        logger.info(f"✅ Group timesheet created: {output_file}")
        return output_file

    def create_individual_timesheets(self) -> List[Path]:
        """Create individual timesheets for each team member"""
        logger.info("👥 Creating individual timesheets...")

        if self.dry_run:
            planned = []
            for member in self.config.team_members:
                output_file = self.output_dir / f"Booked_Hours_{member}.xlsx"
                logger.info(f"[cyan]DRY-RUN[/cyan] Would create individual timesheet → {output_file}")
                planned.append(output_file)
            return planned

        # Load master file
        master_wb = load_workbook(self.master_file_path, data_only=True)
        individual_files = []

        for member in self.config.team_members:
            logger.info(f"  📝 Processing {member}...")

            # Copy template
            output_file = self.output_dir / f"Booked_Hours_{member}.xlsx"
            shutil.copy2(self.individual_template_path, output_file)

            # Load output file
            output_wb = load_workbook(output_file)

            # Transfer data for this member
            self._transfer_individual_data(master_wb, output_wb, member)

            # Save the output file
            output_wb.save(output_file)
            individual_files.append(output_file)
            logger.info(f"    ✅ {member} timesheet created: {output_file}")

        return individual_files

    def _transfer_group_data(self, master_wb, output_wb) -> None:
        """Transfer data from master to group timesheet"""
        try:
            # Transfer PerPerson data (weekly totals)
            if "PerPerson" in master_wb.sheetnames and "PerPerson" in output_wb.sheetnames:
                master_sheet = master_wb["PerPerson"]
                output_sheet = output_wb["PerPerson"]

                # Map team members to their expected row positions in the template
                # Template has names in rows 8-13, we need to find corresponding data in master
                member_row_mapping = {}

                # First, find where each team member is in the master file
                for row in range(8, 20):  # Look for member names starting around row 8
                    name_cell = master_sheet.cell(row=row, column=1)
                    if name_cell.value:
                        master_name = str(name_cell.value).strip()
                        for member in self.config.team_members:
                            if master_name.casefold() == member.casefold():
                                member_row_mapping[member] = row
                                break

                # Now transfer weekly totals to the correct template rows
                # Dynamically find template row positions for each team member
                template_member_rows = {}
                for row in range(8, 20):  # Look for member names in template
                    name_cell = output_sheet.cell(row=row, column=1)

                    # Skip the header row (row 8 with "Name")
                    if row == 8 and name_cell.value and "name" in str(name_cell.value).lower():
                        continue

                    # Map to actual data rows (rows 9-14)
                    if name_cell.value:
                        template_name = str(name_cell.value).strip()
                        for member in self.config.team_members:
                            if template_name.casefold() == member.casefold():
                                template_member_rows[member] = row
                                break

                # Transfer data using dynamic row mapping
                for member in self.config.team_members:
                    template_row = template_member_rows.get(member)
                    master_row = member_row_mapping.get(member)

                    if master_row and template_row:
                        # Transfer weekly totals (columns B to L)
                        for col in range(2, 13):  # Columns B to L (weeks 2-11 + TOTAL)
                            master_cell = master_sheet.cell(row=master_row, column=col)
                            output_cell = output_sheet.cell(row=template_row, column=col)
                            if master_cell.value is not None:
                                output_cell.value = master_cell.value

                logger.info("    ✅ PerPerson data transferred")

            # Transfer Results data
            if "Results" in master_wb.sheetnames and "Results" in output_wb.sheetnames:
                master_sheet = master_wb["Results"]
                output_sheet = output_wb["Results"]

                # Transfer key statistics
                key_cells = [
                    ("B2", "B2"),  # Suggested Total Time
                    ("B3", "B3"),  # Actual Total Time
                    ("B4", "B4"),  # Hours Remaining
                    ("B5", "B5"),  # Hours Remaining from Chosen Budget
                    ("B6", "B6")   # Initial Time Estimate vs Actual
                ]

                for master_cell, output_cell in key_cells:
                    master_value = master_sheet[master_cell].value
                    if master_value is not None:
                        output_sheet[output_cell].value = master_value

                logger.info("    ✅ Results data transferred")

        except Exception as e:
            logger.error(f"    ❌ Error transferring group data: {e}")

    def _transfer_individual_data(self, master_wb, output_wb, member_name: str) -> None:
        """Transfer data for individual team member"""
        try:
            # Update the name in the template
            if "BookedHours" in output_wb.sheetnames:
                output_sheet = output_wb["BookedHours"]

                # Update the title with the member's name
                title_cell = output_sheet.cell(row=1, column=1)
                if title_cell.value:
                    title_cell.value = title_cell.value.replace("<YOUR NAME GOES HERE>", member_name)

                # Transfer data from master file if the member's sheet exists
                # Exact, case-insensitive lookup of member sheet
                member_sheet_name = next(
                    (sn for sn in master_wb.sheetnames if sn.casefold() == member_name.casefold()),
                    None,
                )
                if member_sheet_name:
                    master_sheet = master_wb[member_sheet_name]

                    # Transfer hour records (assuming they start around row 4)
                    for row in range(4, min(100, master_sheet.max_row + 1)):
                        # Check if this row has data
                        week_cell = master_sheet.cell(row=row, column=1)
                        if week_cell.value and str(week_cell.value).strip():
                            # Transfer the row data
                            for col in range(1, min(9, master_sheet.max_column + 1)):
                                master_cell = master_sheet.cell(row=row, column=col)
                                output_cell = output_sheet.cell(row=row, column=col)
                                if master_cell.value is not None:
                                    output_cell.value = master_cell.value

                logger.info(f"    ✅ {member_name} data transferred")
            else:
                logger.warning(f"    ⚠️  BookedHours sheet not found in template")

        except Exception as e:
            logger.error(f"    ❌ Error transferring {member_name} data: {e}")

    def create_zip_file(self, individual_files: List[Path]) -> Optional[Path]:
        """Create a zip file containing all individual timesheets"""
        logger.info("📦 Creating zip file...")

        zip_filename = self.output_dir / f"Booked_Group{self.config.group_number}_Wk{self.week_number}.zip"

        if self.dry_run:
            logger.info(f"[cyan]DRY-RUN[/cyan] Would create zip → {zip_filename}")
            return zip_filename

        try:
            with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file_path in individual_files:
                    # Add file to zip with just the filename (not full path)
                    filename = Path(file_path).name
                    zipf.write(str(file_path), filename)

            logger.info(f"✅ Zip file created: {zip_filename}")
            return zip_filename
        except Exception as e:
            logger.error(f"❌ Error creating zip file: {e}")
            return None

    def create_merged_timesheet(self, individual_files: List[Path]) -> Optional[Path]:
        """Create a single Excel file with all individual timesheets as separate sheets."""
        logger.info("📋 Creating merged timesheet...")

        merged_filename = self.output_dir / f"Booked_Hours_All.xlsx"

        if self.dry_run:
            logger.info(f"[cyan]DRY-RUN[/cyan] Would create merged timesheet → {merged_filename}")
            return merged_filename

        try:
            from openpyxl import Workbook
            merged_wb = Workbook()
            # Remove the default sheet
            merged_wb.remove(merged_wb.active)

            for file_path in individual_files:
                member_name = file_path.stem.replace("Booked_Hours_", "")
                logger.info(f"  📝 Adding {member_name} sheet...")

                # Load the individual timesheet with data_only=True to get values, not formulas
                individual_wb = load_workbook(file_path, data_only=True)
                if "BookedHours" in individual_wb.sheetnames:
                    source_sheet = individual_wb["BookedHours"]

                    # Create new sheet in merged workbook
                    new_sheet = merged_wb.create_sheet(title=member_name)

                    # Copy all data from source sheet
                    for row in source_sheet.iter_rows(values_only=True):
                        new_sheet.append(row)

                    logger.info(f"    ✅ {member_name} sheet added")
                else:
                    logger.warning(f"    ⚠️  BookedHours sheet not found in {file_path}")

            merged_wb.save(merged_filename)
            logger.info(f"✅ Merged timesheet created: {merged_filename}")
            return merged_filename

        except Exception as e:
            logger.error(f"❌ Error creating merged timesheet: {e}")
            return None

    def cleanup_individual_files(self, individual_files: List[Path]) -> None:
        """Remove individual timesheet files after creating merged timesheet."""
        logger.info("🧹 Cleaning up individual timesheet files...")

        if self.dry_run:
            for file_path in individual_files:
                logger.info(f"[cyan]DRY-RUN[/cyan] Would remove: {file_path}")
            return

        try:
            for file_path in individual_files:
                if file_path.exists():
                    file_path.unlink()
                    logger.info(f"  ✅ Removed: {file_path.name}")
                else:
                    logger.warning(f"  ⚠️  File not found: {file_path}")
            logger.info("✅ Individual timesheet files cleaned up")
        except Exception as e:
            logger.error(f"❌ Error cleaning up individual files: {e}")

    def cleanup_master_file(self) -> None:
        """Remove the downloaded master file from the project directory"""
        logger.info("🧹 Cleaning up downloaded master file...")

        if self.dry_run:
            logger.info(f"[cyan]DRY-RUN[/cyan] Would remove: {self.master_file_path}")
            return

        try:
            if self.master_file_path.exists():
                self.master_file_path.unlink()
                logger.info("✅ Master file removed from project directory")
            else:
                logger.warning("⚠️  Master file not found in project directory")
        except Exception as e:
            logger.error(f"❌ Error removing master file: {e}")

    def upload_to_gdrive(self, group_file: Path, zip_file: Path) -> bool:
        """Upload generated files to Google Drive"""
        logger.info("☁️  Uploading files to Google Drive...")

        # Dry-run: only report planned actions
        week_dir = f"Week{self.week_number}"
        gdrive_week_path = f"{self.config.remote_name}:{self.config.remote_base_path}/{week_dir}"

        if self.dry_run:
            logger.info(f"[cyan]DRY-RUN[/cyan] Would run: rclone mkdir {gdrive_week_path}")
            logger.info(f"[cyan]DRY-RUN[/cyan] Would run: rclone copy {group_file} {gdrive_week_path}")
            logger.info(f"[cyan]DRY-RUN[/cyan] Would run: rclone copy {zip_file} {gdrive_week_path}")
            return True

        try:
            # Ensure week directory exists
            mkdir_result = subprocess.run([
                "rclone", "mkdir", gdrive_week_path
            ], capture_output=True, text=True)
            if mkdir_result.returncode != 0:
                logger.error(f"  ❌ Failed to create directory: {mkdir_result.stderr}")
                return False

            # Upload group timesheet
            logger.info("  📤 Uploading group timesheet...")
            group_upload_result = subprocess.run([
                "rclone", "copy", str(group_file), gdrive_week_path
            ], capture_output=True, text=True)
            if group_upload_result.returncode != 0:
                logger.error(f"    ❌ Error uploading group timesheet: {group_upload_result.stderr}")
                return False

            # Upload zip file
            logger.info("  📤 Uploading zip file...")
            zip_upload_result = subprocess.run([
                "rclone", "copy", str(zip_file), gdrive_week_path
            ], capture_output=True, text=True)
            if zip_upload_result.returncode != 0:
                logger.error(f"    ❌ Error uploading zip file: {zip_upload_result.stderr}")
                return False

            logger.info(f"✅ Files uploaded to Google Drive: {gdrive_week_path}")
            return True

        except Exception as e:
            logger.error(f"❌ Error uploading to Google Drive: {e}")
            return False

    def upload_to_remote(self, remote_name: str, remote_base_path: str, group_file: Path, zip_file: Path) -> bool:
        """Upload files to an arbitrary rclone remote under Week<week>."""
        logger.info(f"☁️  Uploading files to {remote_name}:{remote_base_path} ...")
        week_dir = f"Week{self.week_number}"
        remote_week_path = f"{remote_name}:{remote_base_path}/{week_dir}"

        if self.dry_run:
            logger.info(f"[cyan]DRY-RUN[/cyan] Would run: rclone mkdir {remote_week_path}")
            flags = " ".join(self.config.secondary_remote_copy_flags or [])
            logger.info(f"[cyan]DRY-RUN[/cyan] Would run: rclone copy {group_file} {remote_week_path} {flags}")
            logger.info(f"[cyan]DRY-RUN[/cyan] Would run: rclone copy {zip_file} {remote_week_path} {flags}")
            return True

        try:
            mkdir_result = subprocess.run(["rclone", "mkdir", remote_week_path], capture_output=True, text=True)
            if mkdir_result.returncode != 0:
                logger.error(f"  ❌ Failed to create directory: {mkdir_result.stderr}")
                return False

            logger.info("  📤 Uploading group timesheet...")
            up1 = subprocess.run(["rclone", "copy", str(group_file), remote_week_path, * (self.config.secondary_remote_copy_flags or [])], capture_output=True, text=True)
            if up1.returncode != 0:
                logger.error(f"    ❌ Error uploading group timesheet: {up1.stderr}")
                return False

            logger.info("  📤 Uploading zip file...")
            up2 = subprocess.run(["rclone", "copy", str(zip_file), remote_week_path, * (self.config.secondary_remote_copy_flags or [])], capture_output=True, text=True)
            if up2.returncode != 0:
                logger.error(f"    ❌ Error uploading zip file: {up2.stderr}")
                return False

            logger.info(f"✅ Files uploaded to {remote_week_path}")
            return True
        except Exception as e:
            logger.error(f"❌ Error uploading to {remote_name}: {e}")
            return False

    def validate_templates(self) -> bool:
        """Validate that templates exist and contain expected sheet names."""
        ok = True

        if not self.group_template_path.exists():
            logger.error(f"❌ Group template not found: {self.group_template_path}")
            ok = False
        else:
            try:
                wb = load_workbook(self.group_template_path, data_only=True)
                for required_sheet in ("PerPerson", "Results"):
                    if required_sheet not in wb.sheetnames:
                        logger.error(
                            f"❌ Group template missing sheet '{required_sheet}': {self.group_template_path}"
                        )
                        ok = False
            except Exception as e:
                logger.error(f"❌ Failed to open group template: {e}")
                ok = False

        if not self.individual_template_path.exists():
            logger.error(f"❌ Individual template not found: {self.individual_template_path}")
            ok = False
        else:
            try:
                wb = load_workbook(self.individual_template_path, data_only=True)
                if "BookedHours" not in wb.sheetnames:
                    logger.error(
                        f"❌ Individual template missing sheet 'BookedHours': {self.individual_template_path}"
                    )
                    ok = False
            except Exception as e:
                logger.error(f"❌ Failed to open individual template: {e}")
                ok = False

        return ok

    def run(self) -> bool:
        """Main execution function"""
        logger.info("🚀 Starting CITS3200 Spreadsheet Automation")
        logger.info("=" * 50)

        # Clear output directory before beginning
        self.clear_output_dir()

        # Step 0: Validate templates early
        if not self.validate_templates():
            return False

        # Step 1: Prompt for week number if not provided
        if self.week_number is None:
            if not self.prompt_for_week():
                logger.error("❌ Invalid week number. Exiting.")
                return False

        # Step 2: Copy master file from Google Drive
        if not self.copy_master_from_gdrive():
            logger.error("❌ Failed to copy master file. Exiting.")
            return False

        # Step 3: Create group timesheet
        group_file = self.create_group_timesheet()
        if not group_file:
            logger.error("❌ Failed to create group timesheet.")
            return False

        # Step 4: Create individual timesheets
        individual_files = self.create_individual_timesheets()
        if not individual_files:
            logger.error("❌ Failed to create individual timesheets.")
            return False

        # Step 5: Create zip file
        zip_file = self.create_zip_file(individual_files)

        # Step 6: Create merged timesheet
        merged_file = self.create_merged_timesheet(individual_files)

        # Step 7: Clean up individual timesheet files
        self.cleanup_individual_files(individual_files)

        # Step 8: Clean up downloaded master file
        self.cleanup_master_file()

        # Step 9: Prompt for upload to Google Drive
        if group_file and zip_file and merged_file:
            if self.dry_run:
                logger.info("[cyan]DRY-RUN[/cyan] Would prompt for upload to Google Drive; skipping in dry-run.")
                if self.config.secondary_remote_name and self.config.secondary_remote_base_path:
                    logger.info("[cyan]DRY-RUN[/cyan] Would prompt for upload to Teams/SharePoint; skipping in dry-run.")
            else:
                # Prompt for Google Drive upload
                if self.yes_all:
                    response_gdrive = "yes"
                else:
                    try:
                        response_gdrive = input(
                            f"Would you like to upload these to the Week{self.week_number} directory on Google Drive? [y/N] "
                        ).strip().lower()
                    except EOFError:
                        response_gdrive = ""
                if response_gdrive in {"y", "yes"}:
                    if not self.upload_to_gdrive(group_file, zip_file):
                        return False
                else:
                    logger.info("📦 Google Drive upload skipped by user.")

                # Prompt for Teams/SharePoint upload (secondary)
                if self.config.secondary_remote_name and self.config.secondary_remote_base_path:
                    if self.yes_all:
                        response_secondary = "yes"
                    else:
                        try:
                            response_secondary = input(
                                f"Would you also like to upload these to the Week{self.week_number} directory on {self.config.secondary_remote_name}:{self.config.secondary_remote_base_path}? [y/N] "
                            ).strip().lower()
                        except EOFError:
                            response_secondary = ""
                    if response_secondary in {"y", "yes"}:
                        logger.info(f"  ☁️  Uploading files to {self.config.secondary_remote_name}:{self.config.secondary_remote_base_path}...")
                        if not self.upload_to_remote(
                            remote_name=self.config.secondary_remote_name,
                            remote_base_path=self.config.secondary_remote_base_path,
                            group_file=group_file,
                            zip_file=zip_file,
                        ):
                            return False
                        logger.info(f"✅ Files uploaded to {self.config.secondary_remote_name}:{self.config.secondary_remote_base_path}")
                    else:
                        logger.info("📦 Teams/SharePoint upload skipped by user.")

        # Step 10: List created files
        logger.info("\n" + "=" * 50)
        logger.info("✅ Automation completed successfully!")
        logger.info(f"📁 Output files created in: {self.output_dir}")

        if not self.dry_run:
            output_files = sorted(p.name for p in self.output_dir.iterdir())
            logger.info(f"📄 Created {len(output_files)} files:")
            for file in output_files:
                logger.info(f"  - {file}")

        # Optional: prompt to run minutes export
        if not self.dry_run:
            if self.yes_all:
                run_minutes = "yes"
            else:
                try:
                    run_minutes = input("Would you like to export meeting minutes now? [y/N] ").strip().lower()
                except EOFError:
                    run_minutes = ""
            if run_minutes in {"y", "yes"}:
                minutes_script = Path(__file__).with_name("export_minutes.py")
                if minutes_script.exists():
                    logger.info("📝 Launching minutes export...")
                    try:
                        minutes_args = [sys.executable, str(minutes_script)]
                        if self.week_number is not None:
                            minutes_args += ["--week", str(self.week_number)]
                        rc = subprocess.run(minutes_args, check=False)
                        if rc.returncode != 0:
                            logger.error(f"❌ export_minutes.py exited with code {rc.returncode}")
                    except Exception as e:
                        logger.error(f"❌ Failed to run export_minutes.py: {e}")
                else:
                    logger.error("❌ export_minutes.py not found next to this script")

        return True

def _parse_cli() -> Dict:
    """Very small CLI parser to avoid heavy dependencies."""
    import argparse

    parser = argparse.ArgumentParser(description="CITS3200 Spreadsheet Automation")
    parser.add_argument("--week", type=int, help="Week number (2-11)")
    parser.add_argument("--dry-run", action="store_true", help="Print planned actions without writing/uploading")
    parser.add_argument("--yes", action="store_true", help="Answer 'yes' to all prompts (non-interactive mode)")
    parser.add_argument("--config", type=str, help="Path to YAML config file", default=None)
    args = parser.parse_args()
    return {
        "week": args.week,
        "dry_run": args.dry_run,
        "yes": args.yes,
        "config_path": Path(args.config) if args.config else None,
    }


if __name__ == "__main__":
    cli = _parse_cli()
    cfg = load_config(cli["config_path"])
    automation = CITS3200Automation(config=cfg, week_number=cli["week"], dry_run=cli["dry_run"], yes_all=cli["yes"])
    success = automation.run()
    sys.exit(0 if success else 1)
