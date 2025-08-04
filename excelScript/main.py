#!/usr/bin/env python3
"""
CITS3200 Team Leader Spreadsheet Automation
Transfers data from master Google Sheet to template files
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import shutil
from datetime import datetime
import json
import zipfile

class CITS3200Automation:
    def __init__(self):
        self.master_file = "CITS3200_Group_31.xlsx"
        self.group_template = "templates/TimeSheet_GroupX_WkY.xlsx"
        self.individual_template = "templates/Booked_Hours_YourName.xlsx"
        self.output_dir = "output"
        self.team_members = ["Steve", "Ben", "Dongkai", "William", "Jeremy", "Noah"]
        self.group_number = 31
        self.week_number = None

        # Ensure output directory exists
        os.makedirs(self.output_dir, exist_ok=True)

    def prompt_for_week(self):
        """Prompt user for week number"""
        while True:
            try:
                week = int(input("Enter week number (2-11): "))
                if 2 <= week <= 11:
                    self.week_number = week
                    print(f"✅ Week {week} selected")
                    return True
                else:
                    print("❌ Week number must be between 2 and 11")
            except ValueError:
                print("❌ Please enter a valid integer")

    def copy_master_from_gdrive(self):
        """Copy the latest master file from Google Drive"""
        print("📥 Copying master file from Google Drive...")
        try:
            # Use rclone to copy the file
            import subprocess
            result = subprocess.run([
                "rclone", "copy",
                f"gdrive:CITS3200/{self.master_file}",
                "./"
            ], capture_output=True, text=True)

            if result.returncode == 0:
                print("✅ Master file copied successfully")
            else:
                print(f"❌ Error copying master file: {result.stderr}")
                return False
        except Exception as e:
            print(f"❌ Error accessing Google Drive: {e}")
            return False
        return True

    def create_group_timesheet(self):
        """Create the group timesheet from template"""
        print("📊 Creating group timesheet...")

        # Copy template with proper naming
        output_file = os.path.join(self.output_dir, f"TimeSheet_Group{self.group_number}_Wk{self.week_number}.xlsx")
        shutil.copy2(self.group_template, output_file)

        # Load both files
        master_wb = load_workbook(self.master_file, data_only=True)
        output_wb = load_workbook(output_file)

        # Transfer data from master to template
        self._transfer_group_data(master_wb, output_wb)

        # Save the output file
        output_wb.save(output_file)
        print(f"✅ Group timesheet created: {output_file}")
        return output_file

    def create_individual_timesheets(self):
        """Create individual timesheets for each team member"""
        print("👥 Creating individual timesheets...")

        # Load master file
        master_wb = load_workbook(self.master_file, data_only=True)
        individual_files = []

        for member in self.team_members:
            print(f"  📝 Processing {member}...")

            # Copy template
            output_file = os.path.join(self.output_dir, f"Booked_Hours_{member}.xlsx")
            shutil.copy2(self.individual_template, output_file)

            # Load output file
            output_wb = load_workbook(output_file)

            # Transfer data for this member
            self._transfer_individual_data(master_wb, output_wb, member)

            # Save the output file
            output_wb.save(output_file)
            individual_files.append(output_file)
            print(f"    ✅ {member} timesheet created: {output_file}")

        return individual_files

    def _transfer_group_data(self, master_wb, output_wb):
        """Transfer data from master to group timesheet"""
        try:
            # Transfer PerPerson data (weekly totals)
            if "PerPerson" in master_wb.sheetnames and "PerPerson" in output_wb.sheetnames:
                master_sheet = master_wb["PerPerson"]
                output_sheet = output_wb["PerPerson"]

                # Map team members to their expected row positions in the template
                # Template has names in rows 8-13, we need to find corresponding data in master
                member_row_mapping = {}

                # First, find where each team member is in the master file
                for row in range(8, 20):  # Look for member names starting around row 8
                    name_cell = master_sheet.cell(row=row, column=1)
                    if name_cell.value:
                        for member in self.team_members:
                            if member in str(name_cell.value):
                                member_row_mapping[member] = row
                                break

                # Now transfer weekly totals to the correct template rows
                                                                # Dynamically find template row positions for each team member
                template_member_rows = {}
                for row in range(8, 20):  # Look for member names in template
                    name_cell = output_sheet.cell(row=row, column=1)

                    # Skip the header row (row 8 with "Name")
                    if row == 8 and name_cell.value and "name" in str(name_cell.value).lower():
                        continue

                    # Map to actual data rows (rows 9-14)
                    if name_cell.value:
                        for member in self.team_members:
                            if member in str(name_cell.value):
                                template_member_rows[member] = row
                                break

                # Transfer data using dynamic row mapping
                for member in self.team_members:
                    template_row = template_member_rows.get(member)
                    master_row = member_row_mapping.get(member)

                    if master_row and template_row:
                        # Transfer weekly totals (columns B to L)
                        for col in range(2, 13):  # Columns B to L (weeks 2-11 + TOTAL)
                            master_cell = master_sheet.cell(row=master_row, column=col)
                            output_cell = output_sheet.cell(row=template_row, column=col)
                            if master_cell.value is not None:
                                output_cell.value = master_cell.value

                print("    ✅ PerPerson data transferred")

            # Transfer Results data
            if "Results" in master_wb.sheetnames and "Results" in output_wb.sheetnames:
                master_sheet = master_wb["Results"]
                output_sheet = output_wb["Results"]

                # Transfer key statistics
                key_cells = [
                    ("B2", "B2"),  # Suggested Total Time
                    ("B3", "B3"),  # Actual Total Time
                    ("B4", "B4"),  # Hours Remaining
                    ("B5", "B5"),  # Hours Remaining from Chosen Budget
                    ("B6", "B6")   # Initial Time Estimate vs Actual
                ]

                for master_cell, output_cell in key_cells:
                    master_value = master_sheet[master_cell].value
                    if master_value is not None:
                        output_sheet[output_cell].value = master_value

                print("    ✅ Results data transferred")

        except Exception as e:
            print(f"    ❌ Error transferring group data: {e}")

    def _transfer_individual_data(self, master_wb, output_wb, member_name):
        """Transfer data for individual team member"""
        try:
            # Update the name in the template
            if "BookedHours" in output_wb.sheetnames:
                output_sheet = output_wb["BookedHours"]

                # Update the title with the member's name
                title_cell = output_sheet.cell(row=1, column=1)
                if title_cell.value:
                    title_cell.value = title_cell.value.replace("<YOUR NAME GOES HERE>", member_name)

                # Transfer data from master file if the member's sheet exists
                if member_name in master_wb.sheetnames:
                    master_sheet = master_wb[member_name]

                    # Transfer hour records (assuming they start around row 4)
                    for row in range(4, min(100, master_sheet.max_row + 1)):
                        # Check if this row has data
                        week_cell = master_sheet.cell(row=row, column=1)
                        if week_cell.value and str(week_cell.value).strip():
                            # Transfer the row data
                            for col in range(1, min(9, master_sheet.max_column + 1)):
                                master_cell = master_sheet.cell(row=row, column=col)
                                output_cell = output_sheet.cell(row=row, column=col)
                                if master_cell.value is not None:
                                    output_cell.value = master_cell.value

                print(f"    ✅ {member_name} data transferred")
            else:
                print(f"    ⚠️  BookedHours sheet not found in template")

        except Exception as e:
            print(f"    ❌ Error transferring {member_name} data: {e}")

    def create_zip_file(self, individual_files):
        """Create a zip file containing all individual timesheets"""
        print("📦 Creating zip file...")

        zip_filename = os.path.join(self.output_dir, f"Booked_Group{self.group_number}_Wk{self.week_number}.zip")

        try:
            with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file_path in individual_files:
                    # Add file to zip with just the filename (not full path)
                    filename = os.path.basename(file_path)
                    zipf.write(file_path, filename)

            print(f"✅ Zip file created: {zip_filename}")
            return zip_filename
        except Exception as e:
            print(f"❌ Error creating zip file: {e}")
            return None

    def cleanup_master_file(self):
        """Remove the downloaded master file from the project directory"""
        print("🧹 Cleaning up downloaded master file...")

        try:
            if os.path.exists(self.master_file):
                os.remove(self.master_file)
                print("✅ Master file removed from project directory")
            else:
                print("⚠️  Master file not found in project directory")
        except Exception as e:
            print(f"❌ Error removing master file: {e}")

    def upload_to_gdrive(self, group_file, zip_file):
        """Upload generated files to Google Drive"""
        print("☁️  Uploading files to Google Drive...")

        try:
            # Create week directory path
            week_dir = f"Week{self.week_number}"
            gdrive_week_path = f"gdrive:CITS3200/{week_dir}"

            # Check if week directory exists, create if not
            import subprocess
            check_result = subprocess.run([
                "rclone", "ls", gdrive_week_path
            ], capture_output=True, text=True)

            if check_result.returncode != 0:
                print(f"  📁 Creating Week{self.week_number} directory...")
                # Create the directory by copying a dummy file and then removing it
                subprocess.run([
                    "rclone", "copy", "README.md", gdrive_week_path
                ], capture_output=True, text=True)
                subprocess.run([
                    "rclone", "delete", f"{gdrive_week_path}/README.md"
                ], capture_output=True, text=True)
                print(f"  ✅ Week{self.week_number} directory created")
            else:
                print(f"  ✅ Week{self.week_number} directory already exists")

            # Upload group timesheet
            print(f"  📤 Uploading group timesheet...")
            group_upload_result = subprocess.run([
                "rclone", "copy", group_file, gdrive_week_path
            ], capture_output=True, text=True)

            if group_upload_result.returncode == 0:
                print(f"    ✅ Group timesheet uploaded")
            else:
                print(f"    ❌ Error uploading group timesheet: {group_upload_result.stderr}")

            # Upload zip file
            print(f"  📤 Uploading zip file...")
            zip_upload_result = subprocess.run([
                "rclone", "copy", zip_file, gdrive_week_path
            ], capture_output=True, text=True)

            if zip_upload_result.returncode == 0:
                print(f"    ✅ Zip file uploaded")
            else:
                print(f"    ❌ Error uploading zip file: {zip_upload_result.stderr}")

            print(f"✅ Files uploaded to Google Drive: {gdrive_week_path}")
            return True

        except Exception as e:
            print(f"❌ Error uploading to Google Drive: {e}")
            return False

    def run(self):
        """Main execution function"""
        print("🚀 Starting CITS3200 Spreadsheet Automation")
        print("=" * 50)

        # Step 1: Prompt for week number
        if not self.prompt_for_week():
            print("❌ Invalid week number. Exiting.")
            return False

        # Step 2: Copy master file from Google Drive
        if not self.copy_master_from_gdrive():
            print("❌ Failed to copy master file. Exiting.")
            return False

        # Step 3: Create group timesheet
        group_file = self.create_group_timesheet()

        # Step 4: Create individual timesheets
        individual_files = self.create_individual_timesheets()

                # Step 5: Create zip file
        zip_file = self.create_zip_file(individual_files)

        # Step 6: Clean up downloaded master file
        self.cleanup_master_file()

        # Step 7: Upload files to Google Drive
        if group_file and zip_file:
            self.upload_to_gdrive(group_file, zip_file)

        print("\n" + "=" * 50)
        print("✅ Automation completed successfully!")
        print(f"📁 Output files created in: {self.output_dir}")

        # List created files
        output_files = os.listdir(self.output_dir)
        print(f"📄 Created {len(output_files)} files:")
        for file in sorted(output_files):
            print(f"  - {file}")

        return True

if __name__ == "__main__":
    automation = CITS3200Automation()
    automation.run()
